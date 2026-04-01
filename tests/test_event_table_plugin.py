import asyncio
import base64
import json
from pathlib import Path

from openpyxl import load_workbook

from outlook_mail_extractor.models import (
    AttachmentDescriptor,
    EmailDTO,
    PluginExecutionResult,
    PluginExecutionStatus,
)
from outlook_mail_extractor.plugins.event_table import EventTablePlugin


class _FakeActionPort:
    def move_to_folder(self, folder_name: str, create_if_missing: bool = True) -> None:
        del folder_name
        del create_if_missing

    def add_categories(self, categories: list[str]) -> None:
        del categories

    def create_appointment(self, *args, **kwargs) -> None:
        del args
        del kwargs

    def list_attachments(self) -> list[AttachmentDescriptor]:
        return []

    def save_attachment(self, attachment_index: int, destination_path: Path) -> None:
        del attachment_index
        del destination_path


def _build_email_data() -> EmailDTO:
    return EmailDTO(
        subject="原始郵件主旨",
        sender="sender@example.com",
        received="2026-03-18 10:30:00",
        body="",
        tables=[],
        entry_id="00000000123456789ABCDEF",
        store_id="00000000AAAABBBBCCCC",
        internet_message_id="<test-message-id@example.com>",
    )


def test_event_table_writes_excel_row(tmp_path) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(config={"output_file": str(output_file)})

    llm_response = json.dumps(
        {
            "action": "appointment",
            "create": True,
            "subject": "專案會議",
            "start": "2026-03-20T14:00:00",
            "end": "2026-03-20T15:00:00",
            "location": "Teams",
            "body": "討論里程碑",
        }
    )

    result = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )

    assert isinstance(result, PluginExecutionResult)
    assert result.status == PluginExecutionStatus.SUCCESS
    assert result.success is True
    assert output_file.exists()
    assert result.details.get("path") == str(output_file)

    workbook = load_workbook(output_file)
    worksheet = workbook.active
    rows = [row for row in worksheet.iter_rows(values_only=True)]

    assert len(rows) == 2
    header = rows[0]
    assert header == (
        "email_subject",
        "email_sender",
        "email_received",
        "email_entry_id",
        "outlook_open_command",
        "event_subject",
        "start",
        "end",
        "location",
        "body",
        "logged_at",
    )
    row = rows[1]
    assert row[0] == "原始郵件主旨"
    assert row[3] == "00000000123456789ABCDEF"
    assert row[4].startswith("powershell -NoProfile -STA -EncodedCommand ")
    encoded = row[4].split("-EncodedCommand ", 1)[1]
    script = base64.b64decode(encoded).decode("utf-16le")
    assert "$matchSender=$false" in script
    assert "if($delta -gt 1440){continue}" in script
    assert row[5] == "專案會議"
    assert row[6] == "2026-03-20T14:00:00"
    assert row[7] == "2026-03-20T15:00:00"
    assert row[8] == "Teams"
    assert row[9] == "討論里程碑"
    assert row[10]

    assert worksheet.cell(row=2, column=5).hyperlink is None


def test_event_table_open_command_can_enable_sender_match(tmp_path) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(
        config={
            "output_file": str(output_file),
            "open_command_match_sender": True,
        }
    )
    llm_response = '{"action":"appointment","create":true,"subject":"A","start":"2026-03-20T09:00:00","end":"2026-03-20T10:00:00"}'

    result = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )

    assert isinstance(result, PluginExecutionResult)
    assert result.status == PluginExecutionStatus.SUCCESS

    workbook = load_workbook(output_file)
    worksheet = workbook.active
    row = [row for row in worksheet.iter_rows(values_only=True)][1]
    encoded = row[4].split("-EncodedCommand ", 1)[1]
    script = base64.b64decode(encoded).decode("utf-16le")
    assert "$matchSender=$true" in script


def test_event_table_accepts_timezone_datetimes(tmp_path) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(config={"output_file": str(output_file)})

    llm_response = json.dumps(
        {
            "action": "appointment",
            "create": True,
            "subject": "跨時區會議",
            "start": "2026-03-25T12:00:00+08:00",
            "end": "2026-03-25T13:30:00+08:00",
            "location": "Teams",
            "body": "含時區時間",
        }
    )

    result = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )

    assert isinstance(result, PluginExecutionResult)
    assert result.status == PluginExecutionStatus.SUCCESS

    workbook = load_workbook(output_file)
    worksheet = workbook.active
    rows = [row for row in worksheet.iter_rows(values_only=True)]

    assert len(rows) == 2
    row = rows[1]
    assert row[6] == "2026-03-25T12:00:00+08:00"
    assert row[7] == "2026-03-25T13:30:00+08:00"


def test_event_table_noop_when_create_false(tmp_path) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(config={"output_file": str(output_file)})
    llm_response = '{"action":"appointment","create":false}'

    result = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )

    assert isinstance(result, PluginExecutionResult)
    assert result.status == PluginExecutionStatus.SKIPPED
    assert output_file.exists() is False


def test_event_table_appends_without_duplicate_header(tmp_path) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(config={"output_file": str(output_file)})
    llm_response = '{"action":"appointment","create":true,"subject":"A","start":"2026-03-20T09:00:00","end":"2026-03-20T10:00:00"}'

    first = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )
    second = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )

    assert isinstance(first, PluginExecutionResult)
    assert isinstance(second, PluginExecutionResult)
    assert first.status == PluginExecutionStatus.SUCCESS
    assert second.status == PluginExecutionStatus.SUCCESS

    workbook = load_workbook(output_file)
    worksheet = workbook.active
    rows = [row for row in worksheet.iter_rows(values_only=True)]

    assert len(rows) == 3


def test_event_table_ignores_custom_fields_config(tmp_path) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(
        config={
            "output_file": str(output_file),
            "fields": ["custom_a", "custom_b"],
        }
    )
    llm_response = '{"action":"appointment","create":true,"subject":"A","start":"2026-03-20T09:00:00","end":"2026-03-20T10:00:00"}'

    result = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )

    assert isinstance(result, PluginExecutionResult)
    assert result.status == PluginExecutionStatus.SUCCESS

    workbook = load_workbook(output_file)
    worksheet = workbook.active
    rows = [row for row in worksheet.iter_rows(values_only=True)]

    assert rows[0] == (
        "email_subject",
        "email_sender",
        "email_received",
        "email_entry_id",
        "outlook_open_command",
        "event_subject",
        "start",
        "end",
        "location",
        "body",
        "logged_at",
    )


def test_event_table_batch_flush_writes_rows_on_end_job(tmp_path) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(config={"output_file": str(output_file)})
    plugin.begin_job({"batch_flush_enabled": True})

    llm_response = '{"action":"appointment","create":true,"subject":"A","start":"2026-03-20T09:00:00","end":"2026-03-20T10:00:00"}'
    first = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )
    second = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )

    assert first.status == PluginExecutionStatus.SUCCESS
    assert second.status == PluginExecutionStatus.SUCCESS
    assert output_file.exists() is False

    flush_result = plugin.end_job()

    assert isinstance(flush_result, PluginExecutionResult)
    assert flush_result.status == PluginExecutionStatus.SUCCESS
    assert output_file.exists()

    workbook = load_workbook(output_file)
    worksheet = workbook.active
    rows = [row for row in worksheet.iter_rows(values_only=True)]
    assert len(rows) == 3


def test_event_table_retries_when_excel_file_is_locked(tmp_path, monkeypatch) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(
        config={
            "output_file": str(output_file),
            "excel_write_retries": 2,
            "excel_write_retry_delay_seconds": 0,
        }
    )
    llm_response = '{"action":"appointment","create":true,"subject":"A","start":"2026-03-20T09:00:00","end":"2026-03-20T10:00:00"}'

    from openpyxl.workbook.workbook import Workbook

    original_save = Workbook.save
    attempts = {"count": 0}

    def flaky_save(self, filename) -> None:
        attempts["count"] += 1
        if attempts["count"] == 1:
            raise PermissionError("The process cannot access the file")
        original_save(self, filename)

    monkeypatch.setattr(Workbook, "save", flaky_save)

    result = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )

    assert result.status == PluginExecutionStatus.SUCCESS
    assert attempts["count"] == 2
    assert output_file.exists()


def test_event_table_returns_retriable_failed_when_excel_stays_locked(
    tmp_path, monkeypatch
) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(
        config={
            "output_file": str(output_file),
            "excel_write_retries": 1,
            "excel_write_retry_delay_seconds": 0,
        }
    )
    llm_response = '{"action":"appointment","create":true,"subject":"A","start":"2026-03-20T09:00:00","end":"2026-03-20T10:00:00"}'

    from openpyxl.workbook.workbook import Workbook

    def always_locked(self, filename) -> None:
        del filename
        raise PermissionError("The process cannot access the file")

    monkeypatch.setattr(Workbook, "save", always_locked)

    result = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )

    assert result.status == PluginExecutionStatus.RETRIABLE_FAILED
    assert result.code == "excel_file_locked"
    assert "Please close it in Excel and retry" in result.message
    assert output_file.exists() is False


def test_event_table_batch_flush_reports_lock_error(tmp_path, monkeypatch) -> None:
    output_file = tmp_path / "events.xlsx"
    plugin = EventTablePlugin(
        config={
            "output_file": str(output_file),
            "excel_write_retries": 1,
            "excel_write_retry_delay_seconds": 0,
        }
    )
    plugin.begin_job({"batch_flush_enabled": True})
    llm_response = '{"action":"appointment","create":true,"subject":"A","start":"2026-03-20T09:00:00","end":"2026-03-20T10:00:00"}'

    from openpyxl.workbook.workbook import Workbook

    def always_locked(self, filename) -> None:
        del filename
        raise PermissionError("The process cannot access the file")

    monkeypatch.setattr(Workbook, "save", always_locked)

    buffered = asyncio.run(
        plugin.execute(_build_email_data(), llm_response, _FakeActionPort())
    )
    flush_result = plugin.end_job()

    assert buffered.status == PluginExecutionStatus.SUCCESS
    assert isinstance(flush_result, PluginExecutionResult)
    assert flush_result.status == PluginExecutionStatus.RETRIABLE_FAILED
    assert flush_result.code == "excel_file_locked"
    assert "Please close it in Excel and retry" in flush_result.message
