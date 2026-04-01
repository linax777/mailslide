"""Event Table Plugin"""

import base64
from datetime import datetime
from pathlib import Path
from time import sleep

from loguru import logger
from openpyxl import Workbook, load_workbook

from ..models import EmailDTO, MailActionPort, PluginExecutionResult
from .base import BasePlugin, PluginCapability, PluginConfig, register_plugin


class ExcelFileLockedError(Exception):
    """Raised when the target Excel file stays locked after retries."""


@register_plugin
class EventTablePlugin(BasePlugin):
    """Write extracted appointment data to an Excel table."""

    name = "event_table"
    capabilities = {PluginCapability.REQUIRES_LLM}
    default_system_prompt = """你是一個日曆助手。分析以下郵件內容，判斷是否包含預約、會議或行程資訊。

回覆時只輸出 JSON，不要有任何其他文字、解釋或 markdown 格式。"""
    default_response_json_format = {
        "create_true": '{"action": "appointment", "create": true, "subject": "約會主題", "start": "2024-01-15T14:00:00", "end": "2024-01-15T15:00:00", "location": "會議室或線上連結", "body": "額外備註"}',
        "create_false": '{"action": "appointment", "create": false}',
    }
    default_output_file = "output/events.xlsx"
    default_fields = [
        "email_subject",
        "email_sender",
        "email_received",
        "email_entry_id",
        "outlook_open_command",
        "event_subject",
        "start",
        "end",
        "location",
        "body",
        "logged_at",
    ]

    def __init__(self, config: dict | None = None):
        config = config or {}
        self.config = self._load_config(config)
        self.output_file = config.get("output_file", self.default_output_file)
        self.open_command_match_sender = bool(
            config.get("open_command_match_sender", False)
        )
        self.excel_write_retries = self._coerce_non_negative_int(
            config.get("excel_write_retries", 3),
            default=3,
        )
        self.excel_write_retry_delay_seconds = self._coerce_non_negative_int(
            config.get("excel_write_retry_delay_seconds", 1),
            default=1,
        )
        self.fields = list(self.default_fields)
        self._batch_flush_enabled = False
        self._pending_rows: list[dict[str, str]] = []
        if "fields" in config:
            logger.warning(
                "[event_table] 'fields' config is ignored; Excel schema is fixed"
            )

    def begin_job(self, context: dict | None = None) -> None:
        context = context or {}
        self._batch_flush_enabled = bool(context.get("batch_flush_enabled", False))
        self._pending_rows = []

    def end_job(self) -> PluginExecutionResult | None:
        if not self._batch_flush_enabled or not self._pending_rows:
            return None

        try:
            self._append_rows_to_excel(self._pending_rows)
            flushed_count = len(self._pending_rows)
            self._pending_rows = []
            return self.success_result(
                message=f"Flushed {flushed_count} buffered rows to Excel",
                code="batch_flushed",
                details={"flushed_rows": flushed_count},
            )
        except ExcelFileLockedError as error:
            return self.retriable_failed_result(
                message=str(error),
                code="excel_file_locked",
                details={
                    "pending_rows": len(self._pending_rows),
                    "path": str(Path(self.output_file)),
                },
            )
        except Exception as error:
            return self.retriable_failed_result(
                message=f"Batch flush failed: {error}",
                code="batch_flush_failed",
                details={"pending_rows": len(self._pending_rows)},
            )

    def _load_config(self, config: dict) -> PluginConfig:
        return self._load_common_config(
            config,
            response_json_format_default=self.default_response_json_format,
        )

    async def execute(
        self,
        email_data: EmailDTO,
        llm_response: str,
        action_port: MailActionPort,
    ) -> PluginExecutionResult:
        """Write appointment data into Excel when LLM asks to create."""
        del action_port
        try:
            response_data = self._parse_response(llm_response)
            if response_data.get("action") != "appointment":
                return self.skipped_result(
                    message="Action is not appointment",
                    code="action_mismatch",
                )

            if not response_data.get("create", False):
                return self.skipped_result(
                    message="Create flag is false",
                    code="create_false",
                )

            event_subject = response_data.get("subject", "")
            start_str = response_data.get("start", "")
            end_str = response_data.get("end", "")

            if not event_subject or not start_str or not end_str:
                return self.failed_result(
                    message="Missing subject/start/end",
                    code="missing_fields",
                )

            start = self._parse_datetime(start_str)
            end = self._parse_datetime(end_str)
            if not start or not end:
                return self.failed_result(
                    message="Invalid datetime format",
                    code="invalid_datetime",
                )

            entry_id = str(email_data.entry_id).strip()
            store_id = str(email_data.store_id).strip()
            internet_message_id = str(email_data.internet_message_id).strip()
            outlook_open_command = self._build_outlook_open_command(
                entry_id,
                store_id,
                internet_message_id,
                str(email_data.subject),
                str(email_data.received),
                str(email_data.sender),
                self.open_command_match_sender,
            )

            row = {
                "email_subject": str(email_data.subject),
                "email_sender": str(email_data.sender),
                "email_received": str(email_data.received),
                "email_entry_id": entry_id,
                "outlook_open_command": outlook_open_command,
                "event_subject": str(event_subject),
                "start": start.isoformat(timespec="seconds"),
                "end": end.isoformat(timespec="seconds"),
                "location": str(response_data.get("location", "")),
                "body": str(response_data.get("body", "")),
                "logged_at": datetime.now().isoformat(timespec="seconds"),
            }

            if self._batch_flush_enabled:
                self._pending_rows.append(row)
                return self.success_result(message="Event buffered for batch flush")

            self._append_rows_to_excel([row])
            return self.success_result(
                message="Event appended to Excel",
                details={"path": str(Path(self.output_file))},
            )
        except ExcelFileLockedError as error:
            return self.retriable_failed_result(
                message=str(error),
                code="excel_file_locked",
                details={"path": str(Path(self.output_file))},
            )
        except Exception as e:
            return self.retriable_failed_result(
                message=f"Unexpected error: {e}",
                code="unexpected_error",
            )

    def _append_rows_to_excel(self, rows: list[dict[str, str]]) -> None:
        if not rows:
            return

        output_path = Path(self.output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        total_attempts = self.excel_write_retries + 1
        for attempt in range(1, total_attempts + 1):
            workbook = None
            try:
                if output_path.exists() and output_path.stat().st_size > 0:
                    workbook = load_workbook(output_path)
                    worksheet = (
                        workbook["events"]
                        if "events" in workbook.sheetnames
                        else workbook.active
                    )
                else:
                    workbook = Workbook()
                    worksheet = workbook.active
                    worksheet.title = "events"

                if worksheet.max_row == 1 and worksheet.cell(1, 1).value is None:
                    for column_index, field_name in enumerate(self.fields, start=1):
                        worksheet.cell(row=1, column=column_index, value=field_name)

                command_col = self.fields.index("outlook_open_command") + 1
                for row in rows:
                    row_values = [row[field] for field in self.fields]
                    worksheet.append(row_values)
                    row_index = worksheet.max_row
                    worksheet.cell(row=row_index, column=command_col).style = "Normal"

                workbook.save(output_path)
                return
            except (PermissionError, OSError) as error:
                if not self._is_file_lock_error(error):
                    raise

                if attempt >= total_attempts:
                    break

                logger.warning(
                    "[event_table] Excel file is locked: {} "
                    "(retry {}/{} in {} second(s))",
                    output_path,
                    attempt,
                    self.excel_write_retries,
                    self.excel_write_retry_delay_seconds,
                )
                sleep(self.excel_write_retry_delay_seconds)
            finally:
                if workbook is not None:
                    workbook.close()

        raise ExcelFileLockedError(
            "[event_table] Excel file is locked and cannot be written: "
            f"{output_path}. Please close it in Excel and retry."
        )

    @staticmethod
    def _coerce_non_negative_int(value: object, default: int) -> int:
        if isinstance(value, bool):
            return default

        if isinstance(value, int):
            return max(0, value)

        if isinstance(value, str):
            try:
                return max(0, int(value.strip()))
            except ValueError:
                return default

        return default

    @staticmethod
    def _is_file_lock_error(error: PermissionError | OSError) -> bool:
        if isinstance(error, PermissionError):
            return True

        winerror = getattr(error, "winerror", None)
        errno = getattr(error, "errno", None)
        return winerror in {32, 33} or errno == 13

    def _parse_datetime(self, dt_str: str) -> datetime | None:
        """Parse ISO-like datetime string."""
        normalized = dt_str.strip()
        if normalized.endswith("Z"):
            normalized = f"{normalized[:-1]}+00:00"

        try:
            return datetime.fromisoformat(normalized)
        except ValueError:
            pass

        formats = [
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(normalized, fmt)
            except ValueError:
                continue
        return None

    def _build_outlook_open_command(
        self,
        entry_id: str,
        store_id: str = "",
        internet_message_id: str = "",
        email_subject: str = "",
        email_received: str = "",
        email_sender: str = "",
        match_sender: bool = False,
    ) -> str:
        del entry_id
        del store_id
        del internet_message_id
        normalized_subject = email_subject.strip()
        normalized_received = email_received.strip()
        normalized_sender = email_sender.strip()
        if not normalized_subject:
            return ""

        escaped_subject = normalized_subject.replace("'", "''")
        escaped_received = normalized_received.replace("'", "''")
        escaped_sender = normalized_sender.replace("'", "''")
        script = (
            "$ErrorActionPreference='Stop'\n"
            f"$targetSubject='{escaped_subject}'\n"
            f"$targetReceivedRaw='{escaped_received}'\n"
            f"$targetSender='{escaped_sender}'\n"
            f"$matchSender={'$true' if match_sender else '$false'}\n"
            "$subjectCore=($targetSubject -replace '^(?i)(RE|FW|FWD)\\s*:\\s*','').Trim().ToLowerInvariant()\n"
            "$senderCore=$targetSender.Trim().ToLowerInvariant()\n"
            "$app=New-Object -ComObject Outlook.Application\n"
            "$item=$null\n"
            "$targetReceived=$null\n"
            "if(-not [string]::IsNullOrWhiteSpace($targetReceivedRaw)){\n"
            "  try{$targetReceived=[datetime]::Parse($targetReceivedRaw)}catch{}\n"
            "}\n"
            "$stack=New-Object System.Collections.ArrayList\n"
            "foreach($store in $app.Session.Stores){$null=$stack.Add($store.GetRootFolder())}\n"
            "while($stack.Count -gt 0 -and $null -eq $item){\n"
            "  $folder=$stack[$stack.Count-1]\n"
            "  $stack.RemoveAt($stack.Count-1)\n"
            "  try{\n"
            "    $items=$folder.Items\n"
            "    $items.Sort('[ReceivedTime]', $true)\n"
            "    $limit=[Math]::Min($items.Count, 5000)\n"
            "    for($i=1; $i -le $limit -and $null -eq $item; $i++){\n"
            "      $candidate=$items.Item($i)\n"
            "      if($null -eq $candidate){continue}\n"
            "      if($candidate.Class -ne 43){continue}\n"
            "      $candidateSubject=([string]$candidate.Subject).Trim()\n"
            "      if([string]::IsNullOrWhiteSpace($candidateSubject)){continue}\n"
            "      $candidateCore=($candidateSubject -replace '^(?i)(RE|FW|FWD)\\s*:\\s*','').Trim().ToLowerInvariant()\n"
            "      if($candidateCore -ne $subjectCore -and -not $candidateCore.Contains($subjectCore) -and -not $subjectCore.Contains($candidateCore)){continue}\n"
            "      if($null -ne $targetReceived){\n"
            "        $delta=[Math]::Abs((([datetime]$candidate.ReceivedTime)-$targetReceived).TotalMinutes)\n"
            "        if($delta -gt 1440){continue}\n"
            "      }\n"
            "      if($matchSender -and -not [string]::IsNullOrWhiteSpace($senderCore)){\n"
            "        $candidateSender=((([string]$candidate.SenderEmailAddress)+' '+([string]$candidate.SenderName))).ToLowerInvariant()\n"
            "        if(-not [string]::IsNullOrWhiteSpace($candidateSender) -and -not $candidateSender.Contains($senderCore)){continue}\n"
            "      }\n"
            "      $item=$candidate\n"
            "    }\n"
            "  }catch{}\n"
            "  foreach($sub in $folder.Folders){$null=$stack.Add($sub)}\n"
            "}\n"
            'if($null -eq $item){throw "Unable to open message by subject/received time."}\n'
            "$item.Display()\n"
        )
        encoded = base64.b64encode(script.encode("utf-16le")).decode("ascii")
        return f"powershell -NoProfile -STA -EncodedCommand {encoded}"
